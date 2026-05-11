import streamlit as st
import pandas as pd
import warnings
import io
import re
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# ── Machine-independent column width helpers ───────────────────────────────
# Excel stores column widths in "character units" relative to the Normal style
# font's Maximum Digit Width (MDW). MDW varies per machine/OS/DPI, so the same
# character-unit value renders at different pixel widths on different screens.
#
# Fix: we pin MDW=7px (96 DPI, Times New Roman 12pt baseline) and back-calculate
# the character-unit value from our desired pixel target. This makes the stored
# value machine-independent at standard 96 DPI.
#
# Formula:  char_width = (target_pixels - 5_padding) / FIXED_MDW
# openpyxl marks customWidth automatically when .width is assigned.

_FIXED_MDW = 7      # px — Times New Roman 12pt at 96 DPI
_PADDING   = 5      # px — Excel's internal cell padding

def _px_to_char(pixels: float) -> float:
    """Target pixel width → Excel character-unit width (machine-independent)."""
    return (pixels - _PADDING) / _FIXED_MDW

def _apply_pixel_widths(ws, col_pixel_map: dict):
    """
    Set worksheet column widths from a {col_letter: pixels} map.
    Widths are pinned to 96 DPI so they look identical on every machine.
    """
    for col, px in col_pixel_map.items():
        ws.column_dimensions[col].width = _px_to_char(px)
        # Note: openpyxl sets customWidth automatically when .width is assigned

# ── Pixel targets for the 17-column BOC / NonBOC sheet ────────────────────
# Derived from the original character widths:
#   A=4, B=4, C=3, D=12, E=20, F=2, G=12, H=9, I=3, J=4,
#   K=3, L=12, M=20, N=15, O=15, P=6, Q=6
# using pixel = int(char_width * 7 + 5)
_COL17_PIXELS = {
    'A': 33,  'B': 33,  'C': 26,  'D': 89,  'E': 145,
    'F': 19,  'G': 89,  'H': 68,  'I': 26,  'J': 33,
    'K': 26,  'L': 89,  'M': 145, 'N': 110, 'O': 110,
    'P': 47,  'Q': 47,
}

# ── Pixel targets for the NSB sheet (3 columns) ───────────────────────────
# Original widths: A=33.57, B=28.0, C=17.14 char units
# pixel = int(char_width * 7 + 5)
_COL_NSB_PIXELS = {
    'A': 240,   # int(33.57 * 7 + 5)
    'B': 201,   # int(28.0  * 7 + 5)
    'C': 125,   # int(17.14 * 7 + 5)
}

# ── Number format strings for 17-col sheet ────────────────────────────────
COL_FORMATS_17 = {
    'A': '0000', 'B': '0000', 'C': '000', 'D': '000000000000', 'E': 'General',
    'F': '00',   'G': '000000000000', 'H': '000000000', 'I': 'General',
    'J': '0000', 'K': '000', 'L': '000000000000', 'M': 'General',
    'N': 'General', 'O': 'General', 'P': '000000', 'Q': '000000',
}

# ── Bank name normalisation ────────────────────────────────────────────────
BANK_NAME_MAP = {
    "People's Bank": "Peoples Bank",
    "Commercial Bank Of Ceylon PLC": "Commercial Bank PLC",
}
BANK_OF_CEYLON = "Bank of Ceylon"

NSB_ACCOUNT_L   = st.secrets["NSB_ACCOUNT"]
OTHER_ACCOUNT_L = st.secrets["OTHER_ACCOUNT"]

# ── Lookup helpers ─────────────────────────────────────────────────────────
def normalize_bank_name(name):
    if pd.isna(name): return name
    return BANK_NAME_MAP.get(str(name).strip(), str(name).strip())

def lookup_bank_code(branch_df, bank_name):
    bank_name = normalize_bank_name(bank_name)
    if pd.isna(bank_name) or bank_name == "": return ""
    m = branch_df[branch_df["Bank"].str.lower() == bank_name.lower()]
    if m.empty:
        m = branch_df[branch_df["Bank"].str.contains(bank_name, case=False, na=False)]
    return str(m.iloc[0]["Bank Code"]) if not m.empty else "NOT FOUND"

def lookup_branch_code(branch_df, bank_name, branch_name):
    bank_name   = normalize_bank_name(bank_name)
    branch_name = str(branch_name).strip() if not pd.isna(branch_name) else ""
    if pd.isna(bank_name) or bank_name == "" or branch_name == "": return ""
    bank_rows = branch_df[branch_df["Bank"].str.lower() == bank_name.lower()]
    if bank_rows.empty:
        bank_rows = branch_df[branch_df["Bank"].str.contains(bank_name, case=False, na=False)]
    if bank_rows.empty: return "BANK NOT FOUND"
    bm = bank_rows[bank_rows["Branch Name"].str.lower() == branch_name.lower()]
    if bm.empty:
        bm = bank_rows[bank_rows["Branch Name"].str.contains(branch_name, case=False, na=False)]
    return str(bm.iloc[0]["Branch Code"]) if not bm.empty else "BRANCH NOT FOUND"

# ── Excel generators ───────────────────────────────────────────────────────
def generate_rtgs_excel_bytes(df):
    wb = Workbook()
    ws = wb.active
    if len(df) > 0:
        ws.append(list(df.columns))
        for _, row in df.iterrows():
            ws.append(list(row))
        hdr_font = Font(name="Calibri", bold=True, size=11)
        for cell in ws[1]:
            cell.font = hdr_font
    else:
        ws.append(["No RTGS rows in this payment list."])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_nsb_excel_bytes(df):
    wb = Workbook()
    ws = wb.active

    # Machine-independent column widths (pixel-locked at 96 DPI)
    _apply_pixel_widths(ws, _COL_NSB_PIXELS)

    font_hdr  = Font(name="Book Antiqua", size=12, bold=True)
    font_data = Font(name="Book Antiqua", size=12, bold=False)
    font_tot  = Font(name="Book Antiqua", size=12, bold=True)

    align_center = Alignment(horizontal="center")
    align_left   = Alignment(horizontal="left")
    thin         = Side(style="thin")
    border_all   = Border(left=thin, right=thin, top=thin, bottom=thin)

    acc_fmt    = "0"
    amount_fmt = '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* "-"??_);_(@_)'

    def _clean_acc(acc):
        acc = str(acc).strip().replace("-", "").replace(" ", "")
        if len(acc) == 15:   acc = acc[3:]
        elif len(acc) > 12:  acc = acc[-12:]
        return acc

    headers = ["Customer Name", "    Account No. ", "Amount (Rs.)"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font      = font_hdr
        cell.alignment = align_center
        cell.border    = border_all

    for _, row in df.iterrows():
        acc = _clean_acc(row.get("Acc. No", ""))
        try:    acc_num = int(acc)
        except: acc_num = 0
        try:    amount  = float(row.get("Amount", 0))
        except: amount  = 0.0

        ws.append([str(row.get("Customer Name", "")).strip(), acc_num, amount])
        r = ws.max_row
        ws[f"A{r}"].alignment = align_left
        ws[f"B{r}"].alignment = align_center
        ws[f"C{r}"].alignment = align_center
        for col in ["A", "B", "C"]:
            ws[f"{col}{r}"].font   = font_data
            ws[f"{col}{r}"].border = border_all
        ws[f"B{r}"].number_format = acc_fmt
        ws[f"C{r}"].number_format = amount_fmt

    total_row = ws.max_row + 1
    ws[f"C{total_row}"] = f"=SUM(C2:C{total_row - 1})"
    ws[f"C{total_row}"].font          = font_tot
    ws[f"C{total_row}"].alignment     = align_center
    ws[f"C{total_row}"].number_format = amount_fmt

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_17col_excel_bytes(df, account_l):
    def clean_account_number(acc):
        acc = str(acc).strip().replace("-", "").replace(" ", "")
        if len(acc) == 15:   acc = acc[3:]
        elif len(acc) > 12:  acc = acc[-12:]
        return acc

    def to_int_safe(val):
        try:    return int(str(val).strip())
        except: return 0

    def format_amount_int(amount):
        try:    return int(round(float(amount) * 100))
        except: return 0

    def format_maturity_int(mat_date):
        if pd.isna(mat_date): return 0
        if isinstance(mat_date, (datetime.datetime, datetime.date)):
            return int(mat_date.strftime("%y%m%d"))
        try:    return int(pd.to_datetime(mat_date).strftime("%y%m%d"))
        except: return 0

    def format_name(name):
        name = str(name).strip()
        name = re.sub(r'\.([^\s])', r' \1', name)
        return name.replace(".", "")

    wb = Workbook()
    ws = wb.active
    tnr = Font(name="Times New Roman", size=12)

    for _, row in df.iterrows():
        acc_raw = clean_account_number(row.get("Acc. No", ""))
        try:    acc_num = int(acc_raw)
        except: acc_num = 0

        ws.append([
            0,
            to_int_safe(row.get("Bank Code",   0)),
            to_int_safe(row.get("Branch Code", 0)),
            acc_num,
            format_name(row.get("Customer Name", "")),
            23,
            0,
            format_amount_int(row.get("Amount", 0)),
            "slr",
            7010,
            660,
            int(str(account_l).strip()),
            "NSBFMC",
            "NSBFMC",
            "NSBFMC",
            format_maturity_int(row.get("Maturity Date", None)),
            0,
        ])

    # Apply font and number format to every cell
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            cell.font          = tnr
            cell.number_format = COL_FORMATS_17.get(cell.column_letter, "General")

    # Machine-independent column widths (pixel-locked at 96 DPI)
    _apply_pixel_widths(ws, _COL17_PIXELS)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── PRN generator ──────────────────────────────────────────────────────────
def generate_prn_bytes(df, account_l_str, acc_format_fn):
    SPLIT_THRESHOLD = 5_000_000

    def _clean_acc(acc):
        acc = str(acc).strip().replace("-", "").replace(" ", "")
        if len(acc) == 15:   acc = acc[3:]
        elif len(acc) > 12:  acc = acc[-12:]
        return acc

    def _safe_code(val, width):
        v = str(val).strip()
        try:
            int(v)
            return v.zfill(width)
        except ValueError:
            return "0" * width

    def _fmt_amount(amount_float):
        return str(int(round(amount_float * 100))).zfill(9)

    def _split_amounts(amount_float):
        chunks    = []
        remaining = round(amount_float, 2)
        while remaining > SPLIT_THRESHOLD:
            chunks.append(float(SPLIT_THRESHOLD))
            remaining = round(remaining - SPLIT_THRESHOLD, 2)
        chunks.append(remaining)
        return chunks

    def _fmt_maturity(mat):
        if pd.isna(mat): return "000000"
        if isinstance(mat, (datetime.datetime, datetime.date)):
            return mat.strftime("%y%m%d")
        try:    return pd.to_datetime(mat).strftime("%y%m%d")
        except: return "000000"

    def _fmt_name(name):
        name = str(name).strip()
        name = re.sub(r'\.([^\s])', r' \1', name)
        return name.replace(".", "")

    def _build_line(acc, name, bank_code, branch, amount_float, maturity):
        return (
            "0000"                   +
            bank_code                +
            branch                   +
            acc.zfill(12)            +
            name[:20].ljust(20)      +
            "23"                     +
            "000000000000"           +
            _fmt_amount(amount_float)+
            "slr"                    +
            "7010"                   +
            "660"                    +
            account_l_str.zfill(12)  +
            "NSBFMC".ljust(20)       +
            "NSBFMC".ljust(15)       +
            "NSBFMC".ljust(15)       +
            maturity                 +
            "000000"
        )

    lines = []
    for _, row in df.iterrows():
        acc       = acc_format_fn(_clean_acc(row.get("Acc. No", "")))
        name      = _fmt_name(row.get("Customer Name", ""))
        bank_code = _safe_code(row.get("Bank Code",   ""), 4)
        branch    = _safe_code(row.get("Branch Code", ""), 3)
        maturity  = _fmt_maturity(row.get("Maturity Date", None))
        try:    raw_amount = float(row.get("Amount", 0))
        except: raw_amount = 0.0
        for chunk in _split_amounts(raw_amount):
            lines.append(_build_line(acc, name, bank_code, branch, chunk, maturity))

    return "\r\n".join(lines).encode("utf-8")


# ── Streamlit UI ───────────────────────────────────────────────────────────
st.title("Payment List Processor")
st.write("Upload your required files below.")

col1, col2 = st.columns(2)
with col1:
    payment_file   = st.file_uploader("Upload Payment List",   type=["xlsx"])
with col2:
    directory_file = st.file_uploader("Upload Bank Directory", type=["xlsx"])

if st.button("Run Process"):
    if payment_file is not None and directory_file is not None:
        st.write("Processing your files. Please wait.")
        try:
            # Load bank directory
            branch_df = pd.read_excel(directory_file, sheet_name="Branch NEW", header=2)
            branch_df.columns = [
                "Bank", "Bank Code", "Branch Code", "Branch Name",
                "Branch Address", "Tel No1", "Tel No2", "Tel No3", "Fax No", "District"
            ]
            branch_df = branch_df.dropna(subset=["Bank", "Bank Code", "Branch Code"])
            for col in ["Bank", "Branch Name", "Bank Code", "Branch Code"]:
                branch_df[col] = branch_df[col].astype(str).str.strip()

            # Load payment list
            pay_df = pd.read_excel(payment_file, header=1)
            pay_df.columns = pay_df.columns.str.strip()
            pay_df = pay_df.dropna(how="all").reset_index(drop=True)
            pay_df = pay_df[~pay_df.iloc[:, 0].astype(str).str.strip().str.startswith("Print Date")]

            work_df = pay_df[pay_df["Pay By"].astype(str).str.strip() != "Other"].copy()
            work_df["Bank Code"]   = work_df.apply(lambda r: lookup_bank_code(branch_df, r["Bank Name"]), axis=1)
            work_df["Branch Code"] = work_df.apply(lambda r: lookup_branch_code(branch_df, r["Bank Name"], r["Branch Name"]), axis=1)

            # Split by payment type
            rtgs_mask    = work_df["Pay By"].astype(str).str.strip().str.upper() == "RTGS"
            rtgs_df      = work_df[rtgs_mask].copy()
            remaining_df = work_df[~rtgs_mask].copy()

            nsb_mask   = remaining_df["Bank Name"].astype(str).str.strip().str.lower() == "national savings bank"
            nsb_df     = remaining_df[nsb_mask].copy()
            after_nsb  = remaining_df[~nsb_mask].copy()

            boc_mask  = after_nsb["Bank Name"].astype(str).str.strip().str.lower() == BANK_OF_CEYLON.lower()
            boc_df    = after_nsb[boc_mask].copy()
            nonboc_df = after_nsb[~boc_mask].copy()

            st.success("Processing complete. Download your files below.")

            # RTGS
            st.subheader(f"RTGS Files ({len(rtgs_df)} rows)")
            c1, c2 = st.columns(2)
            with c1:
                st.download_button("Download RTGS CSV",   rtgs_df.to_csv(index=False).encode("utf-8"), "RTGS.csv", "text/csv")
            with c2:
                st.download_button("Download RTGS Excel", generate_rtgs_excel_bytes(rtgs_df), "RTGS.xlsx",
                                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            # NSB
            st.subheader(f"NSB Files ({len(nsb_df)} rows)")
            c1, c2, c3 = st.columns(3)
            with c1:
                st.download_button("Download NSB CSV",   nsb_df.to_csv(index=False).encode("utf-8"), "NSB.csv", "text/csv")
            with c2:
                st.download_button("Download NSB Excel", generate_nsb_excel_bytes(nsb_df), "NSB.xlsx",
                                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            with c3:
                st.download_button("Download NSB PRN",   generate_prn_bytes(nsb_df, str(NSB_ACCOUNT_L), lambda acc: acc), "NSB.prn", "text/plain")

            # BOC
            st.subheader(f"BOC Files ({len(boc_df)} rows)")
            c1, c2, c3 = st.columns(3)
            with c1:
                st.download_button("Download BOC CSV",   boc_df.to_csv(index=False).encode("utf-8"), "BOC.csv", "text/csv")
            with c2:
                st.download_button("Download BOC Excel", generate_17col_excel_bytes(boc_df, NSB_ACCOUNT_L), "BOC.xlsx",
                                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            with c3:
                st.download_button("Download BOC PRN",   generate_prn_bytes(boc_df, str(NSB_ACCOUNT_L), lambda acc: acc.zfill(12)), "BOC.prn", "text/plain")

            # Other Banks
            st.subheader(f"Other Banks Files ({len(nonboc_df)} rows)")
            c1, c2, c3 = st.columns(3)
            with c1:
                st.download_button("Download NonBOC CSV",   nonboc_df.to_csv(index=False).encode("utf-8"), "NonBOC.csv", "text/csv")
            with c2:
                st.download_button("Download NonBOC Excel", generate_17col_excel_bytes(nonboc_df, NSB_ACCOUNT_L), "NonBOC.xlsx",
                                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            with c3:
                st.download_button("Download NonBOC PRN",   generate_prn_bytes(nonboc_df, str(NSB_ACCOUNT_L), lambda acc: acc.zfill(12)), "NonBOC.prn", "text/plain")

        except Exception as e:
            st.error(f"An error occurred: {e}")
    else:
        st.warning("Please upload both files before clicking Run Process.")
